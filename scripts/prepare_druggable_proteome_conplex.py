from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


PROTEIN_FIELDS = [
    "protein_id",
    "entry_name",
    "gene_name",
    "protein_name",
    "organism",
    "length",
    "sequence",
    "alphafold_pdb_url",
    "alphafold_cif_url",
    "alphafold_bcif_url",
    "alphafold_global_plddt",
    "alphafold_model_created_date",
    "alphafold_latest_version",
    "source",
    "notes",
    "sequence_key",
    "function_description",
    "target_class",
    "druggable_modalities",
    "max_clinical_phase",
    "disease_icd11_classes",
    "ensembl_id",
]

SEQUENCE_FIELDS = [
    "sequence_key",
    "representative_protein_id",
    "representative_gene_name",
    "representative_protein_name",
    "length",
    "member_count",
    "gene_count",
    "protein_ids",
    "gene_names",
    "sequence",
]

PAIR_FIELDS = ["protein_id", "drug_id", "protein_sequence", "ligand_smiles"]

AA_RE = re.compile(r"^[A-Z*]+$")


def read_drugs(path: Path, smiles_column: str | None = None) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    usable = []
    for row in rows:
        smiles = (
            (row.get(smiles_column) if smiles_column else "")
            or row.get("canonical_smiles")
            or row.get("isomeric_smiles")
            or row.get("smiles")
        )
        if row.get("drug_id") and smiles:
            row["_screening_smiles"] = smiles
            usable.append(row)
    return usable


def normalize_sequence(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace(" ", "").replace("\n", "")


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_proteins(workbook: Path, sheet_name: str) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = [safe_text(value) for value in next(rows_iter)]
    index = {name: idx for idx, name in enumerate(header)}
    required = ["UniProt_ID", "Gene_Symbol", "Protein_Name", "Peptide_Sequence"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    proteins: list[dict[str, str]] = []
    invalid_rows: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    sequence_to_key: dict[str, str] = {}
    sequence_counter = 0

    for row_number, values in enumerate(rows_iter, start=2):
        protein_id = safe_text(values[index["UniProt_ID"]])
        sequence = normalize_sequence(values[index["Peptide_Sequence"]])
        gene = safe_text(values[index["Gene_Symbol"]])
        protein_name = safe_text(values[index["Protein_Name"]])
        if not protein_id or not sequence:
            invalid_rows.append({"row_number": row_number, "protein_id": protein_id, "gene_name": gene, "reason": "missing_id_or_sequence"})
            continue
        if protein_id in seen_ids:
            invalid_rows.append({"row_number": row_number, "protein_id": protein_id, "gene_name": gene, "reason": "duplicate_uniprot_id"})
            continue
        if not AA_RE.match(sequence):
            invalid_rows.append({"row_number": row_number, "protein_id": protein_id, "gene_name": gene, "reason": "non_canonical_sequence_characters"})
            continue
        seen_ids.add(protein_id)
        if sequence not in sequence_to_key:
            sequence_counter += 1
            sequence_to_key[sequence] = f"SEQ{sequence_counter:06d}"
        sequence_key = sequence_to_key[sequence]
        target_class = safe_text(values[index["Target_Class"]]) if "Target_Class" in index else ""
        modalities = safe_text(values[index["Druggable_Modalities"]]) if "Druggable_Modalities" in index else ""
        max_phase = safe_text(values[index["Max_Clinical_Phase"]]) if "Max_Clinical_Phase" in index else ""
        disease_classes = safe_text(values[index["Disease_ICD11_Classes"]]) if "Disease_ICD11_Classes" in index else ""
        function = safe_text(values[index["Function_Description"]]) if "Function_Description" in index else ""
        ensembl = safe_text(values[index["Ensembl_ID"]]) if "Ensembl_ID" in index else ""
        proteins.append(
            {
                "protein_id": protein_id,
                "entry_name": f"{protein_id}_HUMAN",
                "gene_name": gene,
                "protein_name": protein_name,
                "organism": "Homo sapiens (Human)",
                "length": str(len(sequence)),
                "sequence": sequence,
                "alphafold_pdb_url": "",
                "alphafold_cif_url": "",
                "alphafold_bcif_url": "",
                "alphafold_global_plddt": "",
                "alphafold_model_created_date": "",
                "alphafold_latest_version": "",
                "source": "ChEMBL druggable proteome workbook",
                "notes": "druggable proteome target",
                "sequence_key": sequence_key,
                "function_description": function,
                "target_class": target_class,
                "druggable_modalities": modalities,
                "max_clinical_phase": max_phase,
                "disease_icd11_classes": disease_classes,
                "ensembl_id": ensembl,
            }
        )
    return proteins, invalid_rows


def build_sequence_rows(proteins: list[dict[str, str]]) -> list[dict[str, Any]]:
    by_key: dict[str, list[dict[str, str]]] = defaultdict(list)
    for protein in proteins:
        by_key[protein["sequence_key"]].append(protein)

    sequence_rows: list[dict[str, Any]] = []
    for sequence_key in sorted(by_key):
        members = by_key[sequence_key]
        representative = members[0]
        protein_ids = sorted({member["protein_id"] for member in members})
        genes = sorted({member["gene_name"] for member in members if member["gene_name"]})
        sequence_rows.append(
            {
                "sequence_key": sequence_key,
                "representative_protein_id": representative["protein_id"],
                "representative_gene_name": representative["gene_name"],
                "representative_protein_name": representative["protein_name"],
                "length": representative["length"],
                "member_count": len(members),
                "gene_count": len(genes),
                "protein_ids": ";".join(protein_ids),
                "gene_names": ";".join(genes),
                "sequence": representative["sequence"],
            }
        )
    return sequence_rows


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def write_pairs(path: Path, sequence_rows: list[dict[str, Any]], drugs: list[dict[str, str]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle, delimiter="\t")
        for drug in drugs:
            smiles = drug["_screening_smiles"]
            for sequence_row in sequence_rows:
                writer.writerow([sequence_row["sequence_key"], drug["drug_id"], sequence_row["sequence"], smiles])
                count += 1
    return count


def write_run_script(path: Path, pairs_tsv: Path, out_tsv: Path, cache_dir: Path, model_path: Path, device: str, batch_size: int) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = f"""#!/usr/bin/env bash
set -euo pipefail
cd "$(dirname "$0")/../.."
export PYTHONPATH=third_party/ConPLex:${{PYTHONPATH:-}}
mkdir -p {cache_dir.as_posix()}
python -m conplex_dti predict \\
  --data-file {pairs_tsv.as_posix()} \\
  --model-path {model_path.as_posix()} \\
  --outfile {out_tsv.as_posix()} \\
  --data-cache-dir {cache_dir.as_posix()} \\
  --device {device} \\
  --batch-size {batch_size}
"""
    path.write_text(content, encoding="utf-8")
    path.chmod(0o755)


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare the ChEMBL druggable proteome for ConPLex screening.")
    parser.add_argument("--workbook", default="druggable_proteome_chembl(1).xlsx")
    parser.add_argument("--sheet", default="ChEMBL_Targets")
    parser.add_argument("--drugs", default="data/processed/drug_library_pubchem_chembl_mapped.csv")
    parser.add_argument("--smiles-column", default=None)
    parser.add_argument("--out-dir", default="outputs/druggable_proteome")
    parser.add_argument("--model-path", default="third_party/ConPLex/models/BindingDB_ExperimentalValidModel.pt")
    parser.add_argument("--device", default="0")
    parser.add_argument("--batch-size", type=int, default=128)
    args = parser.parse_args()

    workbook = Path(args.workbook)
    drugs_path = Path(args.drugs)
    out_dir = Path(args.out_dir)
    proteins = []
    drugs = read_drugs(drugs_path, args.smiles_column)
    proteins, invalid_rows = read_proteins(workbook, args.sheet)
    sequence_rows = build_sequence_rows(proteins)

    protein_csv = out_dir / "protein_library_druggable_chembl.csv"
    sequence_csv = out_dir / "protein_sequence_representatives.csv"
    pairs_tsv = out_dir / "conplex_pairs_druggable_unique_sequences.tsv"
    predictions_tsv = out_dir / "conplex_predictions_druggable_unique_sequences.tsv"
    cache_dir = out_dir / "conplex_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    run_script = out_dir / "run_conplex_druggable_unique_sequences.sh"
    metadata_json = out_dir / "druggable_proteome_conplex_prep.metadata.json"

    write_csv(protein_csv, proteins, PROTEIN_FIELDS)
    write_csv(sequence_csv, sequence_rows, SEQUENCE_FIELDS)
    pair_count = write_pairs(pairs_tsv, sequence_rows, drugs)
    write_run_script(
        run_script,
        pairs_tsv,
        predictions_tsv,
        cache_dir,
        Path(args.model_path),
        args.device,
        args.batch_size,
    )

    target_class_counts = Counter(row["target_class"] or "NA" for row in proteins)
    duplicate_sequence_members = sum(1 for row in sequence_rows if int(row["member_count"]) > 1)
    metadata = {
        "workbook": str(workbook),
        "sheet": args.sheet,
        "drugs": str(drugs_path),
        "smiles_column": args.smiles_column or "canonical_smiles fallback chain",
        "out_dir": str(out_dir),
        "protein_rows_input_valid": len(proteins),
        "protein_rows_invalid_or_skipped": len(invalid_rows),
        "invalid_or_skipped_examples": invalid_rows[:20],
        "unique_uniprot_ids": len({row["protein_id"] for row in proteins}),
        "unique_gene_symbols": len({row["gene_name"] for row in proteins if row["gene_name"]}),
        "unique_sequences": len(sequence_rows),
        "duplicate_sequence_groups": duplicate_sequence_members,
        "drug_rows_usable": len(drugs),
        "expanded_protein_drug_pairs": len(proteins) * len(drugs),
        "conplex_unique_sequence_pairs": pair_count,
        "target_class_counts": dict(target_class_counts.most_common()),
        "outputs": {
            "protein_csv": str(protein_csv),
            "sequence_csv": str(sequence_csv),
            "pairs_tsv": str(pairs_tsv),
            "predictions_tsv": str(predictions_tsv),
            "run_script": str(run_script),
            "cache_dir": str(cache_dir),
        },
        "run_command": str(run_script),
        "notes": [
            "ConPLex is sequence-based, so duplicate protein sequences are scored once and can be expanded back to all UniProt IDs after prediction.",
            "The pair file uses sequence_key as proteinID for ConPLex prediction.",
        ],
    }
    metadata_json.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(metadata, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
