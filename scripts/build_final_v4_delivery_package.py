#!/usr/bin/env python3
"""Build the audited v4 CSV/XLSX/PDF delivery package after agent review."""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML


def clean(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "null"} else text


def esc(value: Any) -> str:
    return html.escape(clean(value))


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def bool_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(False, index=df.index)
    return df[column].astype(str).str.lower().isin({"true", "1", "1.0", "yes"})


def count_table(series: pd.Series, total: int) -> str:
    rows = []
    for label, count in series.value_counts(dropna=False).items():
        rows.append(
            f"<tr><td>{esc(label or '未标注')}</td><td>{int(count)}</td>"
            f"<td>{100 * int(count) / max(total, 1):.1f}%</td></tr>"
        )
    return "<table><thead><tr><th>类别</th><th>数量</th><th>比例</th></tr></thead><tbody>" + "".join(rows) + "</tbody></table>"


def join_reviewed(formal384: pd.DataFrame, reviewed: pd.DataFrame) -> pd.DataFrame:
    review_columns = [column for column in reviewed.columns if column.startswith("agent_")]
    evidence_columns = [
        column
        for column in reviewed.columns
        if column.startswith("ot_full_")
        or column.startswith("chembl_exact_")
        or column.startswith("chembl_activity_")
        or column.startswith("pair_pubmed_")
        or column.startswith("post_approval_pair_")
        or column.startswith("representative_pair_")
        or column.startswith("representative_post_")
    ]
    extra = reviewed[["pair_id", *review_columns, *evidence_columns]].copy()
    if extra["pair_id"].duplicated().any():
        raise ValueError("Reviewed evidence contains duplicate pair_id values")
    merged = formal384.merge(extra, on="pair_id", how="left", validate="one_to_one")
    if merged["agent_feasibility_grade"].astype(str).str.strip().eq("").any():
        raise ValueError("Agent review is incomplete for one or more final384 rows")
    return merged


def assign_review_queue(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    exact_chembl = out.get("chembl_exact_activity_status", pd.Series("", index=out.index)).astype(str).eq(
        "exact_binding_activity_pchembl_ge_5"
    )
    exact_literature = out["agent_literature_class"].astype(str).eq("exact_pair_validated")
    contradictory = out["agent_literature_class"].astype(str).eq("contradictory")
    grade = out["agent_feasibility_grade"].astype(str)
    confidence = out["agent_confidence"].astype(str)
    resolution = out.get(
        "agent_database_query_resolution", pd.Series("unresolved", index=out.index)
    ).astype(str)
    query_incomplete = (
        ~bool_series(out, "chembl_activity_query_ok") | ~bool_series(out, "lit_ok")
    ) & resolution.ne("resolved_manually")
    out["review_queue"] = "C_manual_review_reserve"
    out.loc[grade.eq("D") | contradictory, "review_queue"] = "D_deprioritize"
    out.loc[grade.isin(["A", "B"]) & ~contradictory, "review_queue"] = "B_novel_testable"
    out.loc[grade.eq("A") & confidence.isin(["high", "medium"]) & ~contradictory, "review_queue"] = "A_novel_priority"
    out.loc[query_incomplete, "review_queue"] = "Q_database_query_incomplete"
    out.loc[exact_chembl | exact_literature, "review_queue"] = "P_validated_control_or_rediscovery"
    queue_order = {
        "A_novel_priority": 0,
        "B_novel_testable": 1,
        "P_validated_control_or_rediscovery": 2,
        "C_manual_review_reserve": 3,
        "Q_database_query_incomplete": 4,
        "D_deprioritize": 5,
    }
    grade_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    confidence_order = {"high": 0, "medium": 1, "low": 2}
    out["_queue_order"] = out["review_queue"].map(queue_order).fillna(9)
    out["_grade_order"] = grade.map(grade_order).fillna(9)
    out["_confidence_order"] = confidence.map(confidence_order).fillna(9)
    out = out.sort_values(
        ["_queue_order", "_grade_order", "_confidence_order", "priority_score_v2", "final384_rank"],
        ascending=[True, True, True, False, True],
        kind="mergesort",
    ).reset_index(drop=True)
    out["review_adjusted_rank"] = range(1, len(out) + 1)
    return out.drop(columns=["_queue_order", "_grade_order", "_confidence_order"])


def chinese_table(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {
        "review_adjusted_rank": "审阅后顺序",
        "final384_rank": "物理筛选顺序",
        "review_queue": "审阅队列",
        "drug_names": "药物",
        "drug_chembl_id": "ChEMBL药物ID",
        "fda_therapeutic_area": "原治疗领域",
        "fda_indication": "原适应症",
        "fda_target_names": "原FDA靶点",
        "primary_gene": "候选新靶点",
        "protein_names": "候选靶点蛋白",
        "target_assay_family_v2": "实验类型",
        "priority_score_v2": "物理优先级分",
        "conplex_score": "ConPLEx分数",
        "boltz_affinity_probability_refined": "Boltz亲和概率",
        "boltz_ligand_iptm_refined": "Boltz配体iPTM",
        "pose_stability_tier": "双样本姿势稳定性",
        "structure_bin": "口袋证据",
        "agent_feasibility_grade": "智能体可行性等级",
        "agent_verdict": "逐条结论",
        "agent_literature_class": "文献核实类别",
        "agent_primary_disease": "优先新适应症假说",
        "agent_repurposing_status": "老药新用类别",
        "agent_disease_evidence": "疾病证据说明",
        "agent_mechanism_rationale": "机制解释",
        "agent_exposure_feasibility": "暴露可行性",
        "agent_active_species_status": "活性物种状态",
        "agent_assay_plan": "建议实验",
        "agent_key_risks": "主要风险",
        "agent_database_query_resolution": "数据库失败补查状态",
        "agent_confidence": "审阅置信度",
        "agent_sources": "核实来源",
        "chembl_exact_activity_status": "ChEMBL精确对状态",
        "chembl_exact_max_binding_pchembl": "ChEMBL最高结合pChEMBL",
        "pair_pubmed_count_2000_2026": "PubMed自动共现数",
        "post_approval_pair_pubmed_count": "上市后自动共现数",
        "default_assay_strategy": "默认实验策略",
        "compound_liability_notes": "化合物风险",
        "brenk_developability_review": "Brenk药化复核",
        "pair_id": "pair_id",
        "is_hot_target_2026": "是否2026会议热点靶点",
        "hot_target_tier": "热点靶点梯队",
        "hot_target_label": "热点靶点名称",
    }
    columns = [column for column in mapping if column in df.columns]
    table = df[columns].rename(columns=mapping).copy()
    for column in ["物理优先级分", "ConPLEx分数", "Boltz亲和概率", "Boltz配体iPTM", "ChEMBL最高结合pChEMBL"]:
        if column in table.columns:
            table[column] = pd.to_numeric(table[column], errors="coerce").round(4)
    return table


def build_assay_matrix(
    final384: pd.DataFrame,
    known_controls: pd.DataFrame,
    mechanisms: pd.DataFrame,
    drug_library: pd.DataFrame,
) -> pd.DataFrame:
    name_by_id = (
        drug_library.drop_duplicates("drug_id").set_index("drug_id")["drug_name"].map(clean).to_dict()
        if {"drug_id", "drug_name"}.issubset(drug_library.columns)
        else {}
    )
    known_by_sequence: dict[str, list[dict[str, str]]] = {}
    for _, row in known_controls.iterrows():
        sequence = clean(row.get("sequence_key"))
        if not sequence:
            continue
        drug_id = clean(row.get("drug_chembl_id"))
        known_by_sequence.setdefault(sequence, []).append(
            {
                "id": drug_id,
                "name": name_by_id.get(drug_id, name_by_id.get(drug_id.split("__")[0], drug_id)),
                "action": clean(row.get("known_action_type")),
                "source": clean(row.get("known_source")) or "known_control_union_v4",
            }
        )
    mechanism_by_gene: dict[str, list[dict[str, str]]] = {}
    if not mechanisms.empty:
        mechanism_scope = mechanisms.copy()
        if "organism" in mechanism_scope.columns:
            mechanism_scope = mechanism_scope[mechanism_scope["organism"].eq("Homo sapiens")]
        if "target_type" in mechanism_scope.columns:
            mechanism_scope = mechanism_scope[mechanism_scope["target_type"].eq("SINGLE PROTEIN")]
        if "direct_interaction" in mechanism_scope.columns:
            mechanism_scope = mechanism_scope[
                mechanism_scope["direct_interaction"].astype(str).str.lower().isin({"true", "1", "1.0"})
            ]
        phase = pd.to_numeric(mechanism_scope.get("max_phase"), errors="coerce").fillna(-1)
        ordered = mechanism_scope.assign(_phase=phase).sort_values(
            ["_phase", "molecule_chembl_id"], ascending=[False, True], kind="mergesort"
        )
        for _, row in ordered.iterrows():
            for gene in re.split(r"[;,|\s]+", clean(row.get("component_gene_symbols"))):
                if not gene:
                    continue
                drug_id = clean(row.get("molecule_chembl_id"))
                mechanism_by_gene.setdefault(gene, []).append(
                    {
                        "id": drug_id,
                        "name": name_by_id.get(drug_id, drug_id),
                        "action": clean(row.get("action_type")),
                        "source": "ChEMBL37_human_single_protein_MoA",
                    }
                )

    rows = []
    for _, row in final384.iterrows():
        controls = known_by_sequence.get(clean(row.get("sequence_key")), [])
        if not controls:
            controls = mechanism_by_gene.get(clean(row.get("primary_gene")), [])
        unique_controls = []
        seen = set()
        for control in controls:
            key = control["id"]
            if key and key not in seen:
                unique_controls.append(control)
                seen.add(key)
            if len(unique_controls) >= 3:
                break
        rows.append(
            {
                "review_adjusted_rank": row.get("review_adjusted_rank", ""),
                "pair_id": row.get("pair_id", ""),
                "drug": row.get("drug_names", ""),
                "candidate_target": row.get("primary_gene", ""),
                "assay_family": row.get("target_assay_family_v2", ""),
                "primary_readout": row.get("default_assay_strategy", ""),
                "agent_specific_assay_plan": row.get("agent_assay_plan", ""),
                "positive_control_ids": ";".join(control["id"] for control in unique_controls),
                "positive_control_names": ";".join(control["name"] for control in unique_controls),
                "positive_control_action_types": ";".join(control["action"] for control in unique_controls),
                "positive_control_sources": ";".join(control["source"] for control in unique_controls),
                "positive_control_review_required": not bool(unique_controls),
                "known_target_counterscreen": row.get("fda_target_names", ""),
                "vehicle_negative_control": "matched vehicle; same solvent concentration",
                "concentration_gate": "solubility/aggregation check plus non-cytotoxic or non-membrane-disruptive concentration window",
                "go_no_go": "dose response in at least two independent repeats plus orthogonal engagement/functional readout; original-target counterscreen must not explain the signal",
                "candidate_action_status": "unknown_requires_functional_assay",
                "compound_risk_review": row.get("compound_liability_notes", ""),
            }
        )
    return pd.DataFrame(rows)


def build_nomination_plate_map(final384: pd.DataFrame) -> pd.DataFrame:
    """Assign 384 pair hypotheses to four 96-position nomination blocks."""

    if len(final384) != 384 or final384["pair_id"].duplicated().any():
        raise ValueError("Nomination plate map requires exactly 384 unique pair rows")
    ordered = final384.sort_values(
        ["target_assay_family_v2", "primary_gene", "review_adjusted_rank"],
        kind="mergesort",
    ).reset_index(drop=True)
    wells = [f"{row}{column:02d}" for row in "ABCDEFGH" for column in range(1, 13)]
    rows = []
    for index, (_, candidate) in enumerate(ordered.iterrows()):
        plate = index // 96 + 1
        well = wells[index % 96]
        rows.append(
            {
                "nomination_block": f"Block_{plate}",
                "position": well,
                "pair_id": candidate["pair_id"],
                "review_adjusted_rank": candidate["review_adjusted_rank"],
                "drug": candidate.get("drug_names", ""),
                "target": candidate.get("primary_gene", ""),
                "assay_family": candidate.get("target_assay_family_v2", ""),
                "review_queue": candidate.get("review_queue", ""),
                "execution_note": (
                    "nomination inventory position only; primary assays, controls and dose series "
                    "must be laid out on target-specific experimental plates"
                ),
            }
        )
    result = pd.DataFrame(rows)
    if not result.groupby("nomination_block").size().eq(96).all():
        raise RuntimeError("Nomination blocks are not exactly 4 x 96")
    return result


def style_excel(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E5F")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column in enumerate(sheet.columns, start=1):
            values = [clean(cell.value) for cell in list(column)[:200]]
            width = min(48, max(10, max((len(value) for value in values), default=10) + 2))
            sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(path)


def compact_rows(df: pd.DataFrame) -> str:
    rows = []
    for _, row in df.iterrows():
        rows.append(
            "<tr>"
            f"<td>{int(row['review_adjusted_rank'])}</td>"
            f"<td>{esc(row.get('drug_names'))}</td>"
            f"<td>{esc(row.get('primary_gene'))}</td>"
            f"<td>{esc(row.get('review_queue'))}</td>"
            f"<td>{esc(row.get('agent_feasibility_grade'))}</td>"
            f"<td>{esc(row.get('agent_primary_disease'))}<br><small>{esc(row.get('agent_repurposing_status'))}</small></td>"
            f"<td>{esc(row.get('agent_verdict'))}</td>"
            "</tr>"
        )
    return "".join(rows)


def build_pdf(
    output: Path,
    final1000: pd.DataFrame,
    final384: pd.DataFrame,
    funnel: dict[str, Any],
    formal: dict[str, Any],
    known: dict[str, Any],
    boltz_known: dict[str, Any] | None = None,
    rank_sensitivity: dict[str, Any] | None = None,
) -> None:
    if "known_calibration" in known:
        known = known["known_calibration"]
    known_id_pairs = int(known.get("known_union_rows", known.get("known_total", 491)))
    known_active_pairs = int(known.get("known_unique_active_moiety_target_rows", known_id_pairs))
    family = final384["target_assay_family_v2"].value_counts()
    queue = final384["review_queue"].value_counts()
    grades = final384["agent_feasibility_grade"].value_counts()
    repurposing = final384["agent_repurposing_status"].value_counts()
    hot_target = bool_series(final384, "is_hot_target_2026")
    exact = final384.get("chembl_exact_activity_status", pd.Series("", index=final384.index)).eq(
        "exact_binding_activity_pchembl_ge_5"
    )
    agent_exact = final384["agent_literature_class"].astype(str).eq("exact_pair_validated")
    validated_union = exact | agent_exact
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    boltz_known = boltz_known or {}
    rank_sensitivity = rank_sensitivity or {}
    if boltz_known:
        known_affinity = boltz_known.get("known_affinity_probability", {})
        discovery_affinity = boltz_known.get("discovery_affinity_probability", {})
        boltz_known_html = (
            "<h3>Boltz 已知阳性校准</h3>"
            f"<p>96 条均衡已知阳性的 affinity probability 中位数为 <b>{float(known_affinity.get('median') or 0):.3f}</b>，"
            f"未标注 discovery Top3000 中位数为 <b>{float(discovery_affinity.get('median') or 0):.3f}</b>；"
            f"条件姿势 A/B 比例分别为 <b>{100*float(boltz_known.get('known_stable_pose_fraction') or 0):.1f}%</b> 与 "
            f"<b>{100*float(boltz_known.get('discovery_stable_pose_fraction') or 0):.1f}%</b>。"
            "该对照只审计已知阳性信号恢复；discovery 未标注，不是负样本，不能计算特异性或准确率。</p>"
        )
    else:
        boltz_known_html = ""
    sections = []
    sections.append(
        f"""
        <section class="cover">
          <div class="kicker">FDA 老药新靶点发现 · 正式 v4</div>
          <h1>全集物理筛选与<br>384 条实验提名</h1>
          <p>ChEMBL-MoA 成药锚点、活性母体统一、ConPLEx、AlphaFold 结构口袋、Boltz-2 双样本精修与逐条文献/可行性审阅</p>
          <div class="stats">
            <div><b>{int(funnel.get('project_physical_pair_rows', 334749)):,}</b><span>唯一结构 × 靶点空间</span></div>
            <div><b>3,000</b><span>Boltz-2 精修</span></div>
            <div><b>1,000</b><span>正式储备池</span></div>
            <div><b>384</b><span>实验提名 pair</span></div>
          </div>
          <div class="warning">本包是药物-候选靶点互作假说及实验优先级，不等同于已证实结合、作用方向或疾病疗效。</div>
          <small>生成时间：{generated}</small>
        </section>
        """
    )
    recall_id = known.get(
        "recall_project_target_universe_by_drug_rank",
        known.get(
            "recall_project462_by_drug_rank",
            known.get("recall_by_drug_rank", known.get("recall", {})),
        ),
    )
    recall = known.get("active_moiety_collapsed_metrics", {}).get(
        "recall_project_target_universe_by_drug_rank", recall_id
    )
    sections.append(
        f"""
        <section>
          <h2>1. 冻结口径与筛选漏斗</h2>
          <table><thead><tr><th>阶段</th><th>数量</th><th>定义</th></tr></thead><tbody>
            <tr><td>FDA 药物原始库</td><td>{int(funnel.get('drug_library_rows', 915)):,}</td><td>保留原始记录用于审计</td></tr>
            <tr><td>项目药物记录</td><td>{int(funnel.get('project_drug_rows', funnel.get('project_drugs', 750))):,}</td><td>FDA/ChEMBL ID 口径；去盐后为 {int(funnel.get('project_unique_model_ligands', 723)):,} 个唯一模型配体结构</td></tr>
            <tr><td>直接小分子靶点</td><td>{int(funnel.get('project_target_rows', funnel.get('project_targets', 463))):,}</td><td>非 GPCR，具 ChEMBL-MoA/OT 可做性和结构口袋支持</td></tr>
            <tr><td>ID 记录空间</td><td>{int(funnel.get('project_cartesian_rows', 347250)):,}</td><td>750 × 463；保留盐型/产品 ID 审计</td></tr>
            <tr><td>唯一物理 pair 空间</td><td>{int(funnel.get('project_physical_pair_rows', 334749)):,}</td><td>{int(funnel.get('project_unique_model_ligands', 723)):,} 个模型结构 × 463；用于靶点内 rank 与全局百分位</td></tr>
            <tr><td>结构与规则 eligible</td><td>{int(funnel.get('eligible_rows', funnel.get('physics_eligible_rows', 0))):,}</td><td>已知对、同源/同家族泄漏和严重化合物风险排除</td></tr>
            <tr><td>Boltz-2 精修</td><td>{int(formal.get('top3000_rows', 3000)):,}</td><td>严格多样性上限下的 Top3000</td></tr>
            <tr><td>正式储备池</td><td>{len(final1000):,}</td><td>完整 Boltz 输出后连续证据排序</td></tr>
            <tr><td>逐条审阅池</td><td>{int(formal.get('agent_review_pool_rows', len(final384))):,}</td><td>口袋 A/B、姿势稳定、非离子通道；用于审阅后替换</td></tr>
            <tr><td>实验提名</td><td>{len(final384):,}</td><td>排除 D/矛盾项后按硬性多样性上限回填</td></tr>
          </tbody></table>
          <h3>已知对校准</h3>
          <p>共有 {known_id_pairs} 条已知 direct-action ID-pair，按模型配体结构折叠后为 {known_active_pairs} 条；仅用于检查漏斗是否过度杀伤，未进入 discovery 候选。以 {known_active_pairs} 条结构折叠对为主口径，ConPLEx 药物内召回：Top10 {100*float(recall.get('top10', recall.get('10', 0))):.1f}%、Top50 {100*float(recall.get('top50', recall.get('50', 0))):.1f}%、Top100 {100*float(recall.get('top100', recall.get('100', 0))):.1f}%、Top300 {100*float(recall.get('top300', recall.get('300', 0))):.1f}%；ID 口径 Top100/Top300 为 {100*float(recall_id.get('top100', recall_id.get('100', 0))):.1f}% / {100*float(recall_id.get('top300', recall_id.get('300', 0))):.1f}%。这些数字可能受训练知识重叠影响，不是无泄漏泛化准确率。</p>
          <p>结构序列审计发现 Top3000 中有 <b>{int(formal.get('structure_sequence_mismatch_rows', 140))}</b> 条的 ConPLEx 蛋白序列与实际 Boltz/P2Rank 模板不一致；这些条目保留在全量审计表，但已从 final1000、审阅池和 final384 硬隔离。</p>
          {boltz_known_html}
        </section>
        """
    )
    sections.append(
        f"""
        <section>
          <h2>2. 证据模块与计分</h2>
          <table><thead><tr><th>模块</th><th>输入</th><th>正式用途</th><th>限制</th></tr></thead><tbody>
            <tr><td>ConPLEx（25）</td><td>统一配体 SMILES + 蛋白序列</td><td>全局、药物内和靶点内相对排序的单一复合百分位</td><td>不是 Kd/Ki；训练知识可能与已知药理重叠</td></tr>
            <tr><td>Boltz-2（30）</td><td>AlphaFold 模板、口袋约束、配体</td><td>亲和概率 24 分 + 双样本条件姿势稳定 6 分</td><td>计算结构证据，不是实验结合事实</td></tr>
            <tr><td>口袋共识（15）</td><td>P2Rank + PUResNet</td><td>判断是否存在可解释、可实验的结合口袋</td><td>靶点可成药不等于任意药物可结合</td></tr>
            <tr><td>OT 可做性（10）</td><td>Open Targets 26.06 tractability</td><td>结构、小分子和临床可做性校正</td><td>不作疾病疗效分</td></tr>
            <tr><td>药物/实验（15）</td><td>物性、干扰警报、assay family</td><td>暴露和实验可执行性优先</td><td>仍需实测浓度窗口</td></tr>
            <tr><td>新颖性（5）</td><td>已知靶点、同源和同家族审计</td><td>阻止已知/家族扩展被包装为新发现</td><td>文献排重仍需人工核实</td></tr>
          </tbody></table>
          <p>该计分是复合排序特征，不应解释为相互独立证据的统计相乘：ConPLEx 的全局/药物内/靶点内排名来自同一模型，口袋、tractability 与实验可行性也共享部分结构先验。药物 ID 加权与 723 个唯一结构折叠两种预筛 Top3000 重叠 {100*float(rank_sensitivity.get('overlap_fraction', 0.998)):.1f}%（{int(rank_sensitivity.get('overlap_rows', 2994)):,}/3,000）；正式终分使用结构折叠 rank。疾病证据位于物理筛选之后：Open Targets 提供 target-disease 支持，智能体只为可检验候选提出一个优先病种；未知作用方向明确保留为待实验变量。</p>
        </section>
        """
    )
    sections.append(
        """
        <section>
          <h2>3. 方法来源与证据边界</h2>
          <ul>
            <li>ConPLEx：Singh 等，<i>Contrastive learning in protein language space predicts interactions between drugs and protein targets</i>，PMID 37289807。<a href="https://pubmed.ncbi.nlm.nih.gov/37289807/">PubMed</a></li>
            <li>Boltz-2：Wohlwend 等，<i>Boltz-2: Towards Accurate and Efficient Binding Affinity Prediction</i>，PMID 40667369 / PMCID PMC12262699。<a href="https://pmc.ncbi.nlm.nih.gov/articles/PMC12262699/">全文</a></li>
            <li>P2Rank：Krivák 与 Hoksza，<i>P2Rank: machine learning based tool for rapid and accurate prediction of ligand binding sites from protein structure</i>。<a href="https://pmc.ncbi.nlm.nih.gov/articles/PMC6091426/">全文</a></li>
            <li>PUResNet：Kandel 等，<i>PUResNet: prediction of protein-ligand binding sites using deep residual neural network</i>。<a href="https://pmc.ncbi.nlm.nih.gov/articles/PMC8424938/">全文</a></li>
            <li>Open Targets tractability：小分子结构、配体、口袋与临床先例字段定义。<a href="https://platform-docs.opentargets.org/target/tractability">官方文档</a></li>
            <li>模型偏倚警示：Graber 等，<i>Resolving data bias improves generalization in binding affinity prediction</i>。<a href="https://www.nature.com/articles/s42256-025-01124-5">Nature Machine Intelligence</a></li>
          </ul>
          <p>这些文献支持各计算模块的用途，不证明本项目任一具体 drug-target pair。已知对召回可能受训练集重叠影响；口袋约束下的结构复现是条件证据；疾病关联不提供作用方向。最终结论仍由直接结合/功能实验决定。</p>
        </section>
        """
    )
    sections.append(
        f"""
        <section>
          <h2>4. final384 质量审阅</h2>
          <div class="twocol"><div><h3>审阅队列</h3>{count_table(queue, len(final384))}</div><div><h3>可行性等级</h3>{count_table(grades, len(final384))}</div></div>
          <div class="twocol"><div><h3>老药新用类别</h3>{count_table(repurposing, len(final384))}</div><div><h3>实验类型</h3>{count_table(family, len(final384))}</div></div>
          <p>命中用户提供的 AACR/ASCO 2026 热点靶点清单：<b>{int(hot_target.sum())}</b> / {len(final384)}。该字段不参与筛选分，仅用于覆盖审计；抗体主导的表面靶点仍受本项目 direct-small-molecule 范围限制。</p>
          <div><h3>精确对排重</h3><p>ChEMBL 中存在 pChEMBL ≥ 5 的精确结合记录：<b>{int(exact.sum())}</b> / {len(final384)}；人工文献审阅判定精确 pair 已验证：<b>{int(agent_exact.sum())}</b> / {len(final384)}；两者并集：<b>{int(validated_union.sum())}</b> / {len(final384)}（{100*int(validated_union.sum())/len(final384):.1f}%）。这些条目保留为已验证对照或再发现，不作为新靶点发现包装；正式包设置最多 96 条的 control/rediscovery 硬配额，不强行凑比例。</p><p>审阅池中每一条均给出文献类别、暴露判断、优先病种、机制解释、实验方案和主要风险；最终 384 为审阅后的可执行子集。</p></div>
        </section>
        """
    )
    sections.append(
        f"""
        <section>
          <h2>5. 实验执行原则</h2>
          <ol>
            <li>交付表提供四个 96 位候选清单块，用于样品管理；真正实验必须按靶点和 assay family 重排，并预留阳性/阴性对照与剂量梯度。</li>
            <li>先确定药物非毒性、非聚集和非膜干扰浓度窗口，再做 purified target 或 target-engagement assay。</li>
            <li>每组必须有已知配体阳性对照、vehicle 阴性对照和药物原已知靶点 counterscreen。</li>
            <li>命中标准至少包含剂量依赖、独立重复和正交 readout；单点信号不进入疾病机制结论。</li>
            <li>功能方向（抑制、激动、拮抗或调节）在结合命中后单独测定，再决定具体疾病。</li>
          </ol>
          <div class="warning">优先顺序是资源配置建议。任何候选都必须经过化合物身份、溶解度、暴露和 assay interference 复核。</div>
        </section>
        """
    )
    sections.append(
        f"""
        <section class="appendix landscape">
          <h2>附录：384 条逐条审阅摘要</h2>
          <table class="compact"><thead><tr><th>#</th><th>药物</th><th>靶点</th><th>队列</th><th>级别</th><th>优先病种假说</th><th>逐条结论</th></tr></thead><tbody>{compact_rows(final384)}</tbody></table>
        </section>
        """
    )
    css = """
    @font-face { font-family: NotoCJK; src: url(file:///usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc); }
    @font-face { font-family: NotoCJK; src: url(file:///usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc); font-weight: 700; }
    @page { size: A4; margin: 14mm 13mm 15mm; @bottom-center { content: counter(page); font: 8pt NotoCJK; color: #667; } }
    @page landscape { size: A4 landscape; margin: 10mm; }
    * { box-sizing: border-box; }
    body { font: 9.5pt/1.55 NotoCJK, sans-serif; color: #1c2730; margin: 0; }
    section { page-break-before: always; }
    section:first-child { page-break-before: auto; }
    .cover { padding-top: 35mm; min-height: 240mm; }
    .kicker { color: #16716b; font-weight: 700; letter-spacing: 0; }
    h1 { font-size: 30pt; line-height: 1.2; margin: 10mm 0 6mm; color: #153945; }
    h2 { font-size: 18pt; color: #153945; border-bottom: 2px solid #2a827a; padding-bottom: 3mm; }
    h3 { font-size: 12pt; color: #1f5e65; margin: 6mm 0 2mm; }
    .stats { display: grid; grid-template-columns: repeat(4, 1fr); gap: 3mm; margin: 14mm 0; }
    .stats div { border-top: 3px solid #2a827a; padding: 4mm 2mm; background: #f3f7f6; }
    .stats b { display: block; font-size: 18pt; color: #153945; }
    .stats span { font-size: 8pt; color: #566; }
    .warning { background: #fff4df; border-left: 4px solid #d88a15; padding: 4mm; margin: 6mm 0; }
    table { width: 100%; border-collapse: collapse; margin: 3mm 0 5mm; font-size: 8.3pt; }
    th { background: #1f5963; color: white; font-weight: 700; text-align: left; }
    th, td { border: .4px solid #cbd5d6; padding: 2mm; vertical-align: top; }
    tr:nth-child(even) td { background: #f6f8f8; }
    .twocol { display: grid; grid-template-columns: 1fr 1fr; gap: 6mm; }
    .compact { font-size: 6.5pt; line-height: 1.25; table-layout: fixed; }
    .compact th, .compact td { padding: 1mm; overflow-wrap: anywhere; }
    .compact th:nth-child(1) { width: 4%; } .compact th:nth-child(2) { width: 14%; }
    .compact th:nth-child(3) { width: 8%; } .compact th:nth-child(4) { width: 16%; }
    .compact th:nth-child(5) { width: 5%; } .compact th:nth-child(6) { width: 17%; }
    .compact th:nth-child(7) { width: 36%; }
    .appendix { page: landscape; }
    .appendix table { page-break-inside: auto; }
    .appendix tr { page-break-inside: avoid; }
    """
    document = "<!doctype html><html lang='zh-CN'><head><meta charset='utf-8'><style>" + css + "</style></head><body>" + "".join(sections) + "</body></html>"
    output.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=document, base_url=str(output.parent)).write_pdf(output)


def build_detailed_cards_pdf(output: Path, final384: pd.DataFrame) -> None:
    pages = []
    ordered = final384.sort_values("review_adjusted_rank", kind="mergesort")
    for start in range(0, len(ordered), 4):
        cards = []
        for _, row in ordered.iloc[start : start + 4].iterrows():
            sources = esc(row.get("agent_sources"))
            cards.append(
                f"""
                <article class="card">
                  <header><b>#{int(row['review_adjusted_rank'])} {esc(row.get('drug_names'))} → {esc(row.get('primary_gene'))}</b><span>{esc(row.get('review_queue'))} · {esc(row.get('agent_feasibility_grade'))}</span></header>
                  <div class="meta"><b>原适应症</b> {esc(row.get('fda_indication')) or '未记录'}<br><b>原 FDA 靶点</b> {esc(row.get('fda_target_names')) or '未记录'}<br><b>候选蛋白</b> {esc(row.get('protein_names'))}<br><b>物理分</b> {esc(round(float(row.get('priority_score_v2') or 0), 2))}；ConPLEx {esc(round(float(row.get('conplex_score') or 0), 4))}；Boltz affinity {esc(round(float(row.get('boltz_affinity_probability_refined') or 0), 4))}<br><b>口袋/姿势</b> {esc(row.get('structure_bin'))}；{esc(row.get('pose_stability_tier'))}</div>
                  <dl>
                    <dt>逐条结论</dt><dd>{esc(row.get('agent_verdict'))}</dd>
                    <dt>文献核实</dt><dd>{esc(row.get('agent_literature_class'))}；ChEMBL：{esc(row.get('chembl_exact_activity_status')) or '无记录'}</dd>
                    <dt>优先病种</dt><dd>{esc(row.get('agent_primary_disease'))}</dd>
                    <dt>老药新用类别</dt><dd>{esc(row.get('agent_repurposing_status'))}</dd>
                    <dt>疾病证据</dt><dd>{esc(row.get('agent_disease_evidence'))}</dd>
                    <dt>机制假说</dt><dd>{esc(row.get('agent_mechanism_rationale'))}</dd>
                    <dt>暴露可行性</dt><dd>{esc(row.get('agent_exposure_feasibility'))}</dd>
                    <dt>活性物种</dt><dd>{esc(row.get('agent_active_species_status'))}</dd>
                    <dt>建议实验</dt><dd>{esc(row.get('agent_assay_plan'))}</dd>
                    <dt>主要风险</dt><dd>{esc(row.get('agent_key_risks'))}</dd>
                    <dt>来源</dt><dd class="sources">{sources}</dd>
                  </dl>
                </article>
                """
            )
        pages.append(f"<section class='page'>{''.join(cards)}</section>")
    css = """
    @font-face { font-family: NotoCJK; src: url(file:///usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc); }
    @font-face { font-family: NotoCJK; src: url(file:///usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc); font-weight: 700; }
    @page { size: A4 landscape; margin: 8mm; @bottom-center { content: counter(page); font: 7pt NotoCJK; color: #667; } }
    * { box-sizing: border-box; }
    body { margin: 0; color: #1c2730; font: 7.2pt/1.35 NotoCJK, sans-serif; }
    .page { display: grid; grid-template-columns: 1fr 1fr; grid-template-rows: 1fr 1fr; gap: 4mm; height: 194mm; page-break-after: always; }
    .page:last-child { page-break-after: auto; }
    .card { border: .6px solid #aebdc0; padding: 3mm; overflow: hidden; }
    header { display: flex; justify-content: space-between; gap: 3mm; border-bottom: 2px solid #2a827a; padding-bottom: 1.5mm; margin-bottom: 1.5mm; color: #153945; }
    header b { font-size: 9pt; } header span { font-size: 6.5pt; text-align: right; }
    .meta { background: #f3f7f6; padding: 1.5mm; margin-bottom: 1.5mm; }
    dl { display: grid; grid-template-columns: 16mm 1fr; margin: 0; }
    dt { font-weight: 700; color: #1f5e65; padding: .7mm 1mm .7mm 0; border-bottom: .3px solid #dce3e4; }
    dd { margin: 0; padding: .7mm 0; border-bottom: .3px solid #dce3e4; overflow-wrap: anywhere; }
    .sources { font-size: 6.2pt; }
    """
    document = "<!doctype html><html lang='zh-CN'><head><meta charset='utf-8'><style>" + css + "</style></head><body>" + "".join(pages) + "</body></html>"
    output.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=document, base_url=str(output.parent)).write_pdf(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--formal-dir", required=True)
    parser.add_argument("--reviewed384", default="")
    parser.add_argument(
        "--selected384",
        required=True,
        help="Post-review selected table. When present it replaces the pre-review nomination.",
    )
    parser.add_argument(
        "--allow-legacy-nonformal-delivery",
        action="store_true",
        help="Deprecated diagnostic path only; never labels the output formal.",
    )
    parser.add_argument("--opentargets-final1000", required=True)
    parser.add_argument("--funnel-summary", required=True)
    parser.add_argument("--known-summary", required=True)
    parser.add_argument("--boltz-known-summary", required=True)
    parser.add_argument("--rank-sensitivity-summary", required=True)
    parser.add_argument(
        "--known-controls",
        default="outputs/current_production_package_v2/full_untruncated_universe_v4/known_control_union_v4.csv",
    )
    parser.add_argument(
        "--mechanisms",
        default="outputs/target_catalog_quality_audit_v1/chembl37_mechanisms_with_human_target_map.csv",
    )
    parser.add_argument("--drug-library", default="data/processed/drug_library_active_moiety_v4.csv")
    parser.add_argument("--config", default="configs/current_pipeline_v4.yaml")
    parser.add_argument("--hot-targets", default="configs/aacr_asco_2026_hot_targets.csv")
    parser.add_argument("--chembl-cache", required=True)
    parser.add_argument("--pubmed-cache", required=True)
    parser.add_argument(
        "--sequence-audit-summary",
        default=(
            "outputs/current_production_package_v2/full_untruncated_universe_v4/"
            "target_sequence_integrity_v4/target_sequence_integrity_v4.summary.json"
        ),
    )
    parser.add_argument("--out-dir", required=True)
    args = parser.parse_args()

    formal_dir = Path(args.formal_dir)
    final1000_path = formal_dir / "final1000_candidates_v4_complete.csv"
    final384_path = formal_dir / "final384_nomination_v4_complete.csv"
    formal_summary_path = formal_dir / "formal_completion_summary_v4_complete.json"
    required_paths = [final1000_path, formal_summary_path]
    required_paths.append(Path(args.selected384))
    for path in required_paths:
        if not path.exists():
            raise FileNotFoundError(path)
    final1000 = pd.read_csv(final1000_path, low_memory=False).fillna("")
    formal_summary = read_json(formal_summary_path)
    if sha256(final1000_path) != str(formal_summary.get("output_sha256", {}).get("final1000", "")):
        raise RuntimeError("Formal final1000 SHA-256 does not match its completion summary")
    if args.opentargets_final1000:
        ot = pd.read_csv(args.opentargets_final1000, low_memory=False).fillna("")
        ot_columns = [
            column
            for column in ot.columns
            if column.startswith("ot_full_") and "known_drug" not in column
        ]
        final1000 = final1000.merge(
            ot[["pair_id", *ot_columns]].drop_duplicates("pair_id"), on="pair_id", how="left", validate="one_to_one"
        )
    if args.allow_legacy_nonformal_delivery:
        raise ValueError(
            "Legacy nonformal delivery is no longer implemented in the v4 formal builder; "
            "use an archived historical script for diagnostics"
        )
    selected = pd.read_csv(args.selected384, low_memory=False).fillna("")
    selected_summary_path = Path(args.selected384).with_suffix(".summary.json")
    if not selected_summary_path.exists():
        raise FileNotFoundError(selected_summary_path)
    selected_summary = read_json(selected_summary_path)
    if sha256(Path(args.selected384)) != str(
        selected_summary.get("output_sha256", {}).get("selected_final384", "")
    ):
        raise RuntimeError("Post-review final384 SHA-256 does not match its selection summary")
    final384 = assign_review_queue(selected)
    hot_targets = pd.read_csv(args.hot_targets, low_memory=False).fillna("")
    if hot_targets["gene"].duplicated().any():
        raise ValueError("Hot-target annotation contains duplicate genes")
    hot_targets["is_hot_target_2026"] = True
    hot_columns = ["gene", "is_hot_target_2026", "hot_target_tier", "hot_target_label", "modality_note"]
    final1000 = final1000.merge(
        hot_targets[hot_columns].rename(columns={"gene": "primary_gene"}),
        on="primary_gene",
        how="left",
        validate="many_to_one",
    )
    final384 = final384.merge(
        hot_targets[hot_columns].rename(columns={"gene": "primary_gene"}),
        on="primary_gene",
        how="left",
        validate="many_to_one",
    )
    final1000["is_hot_target_2026"] = bool_series(final1000, "is_hot_target_2026")
    final384["is_hot_target_2026"] = bool_series(final384, "is_hot_target_2026")
    if len(final1000) != 1000 or len(final384) != 384:
        raise ValueError(f"Formal row contracts failed: final1000={len(final1000)}, final384={len(final384)}")
    if not set(final384["pair_id"]).issubset(set(final1000["pair_id"])):
        raise ValueError("Final384 is not a subset of the formal final1000 reserve")
    forbidden_queues = {"D_deprioritize", "Q_database_query_incomplete"}
    observed_forbidden = sorted(set(final384["review_queue"]) & forbidden_queues)
    contradictory_rows = int(
        final384["agent_literature_class"].astype(str).eq("contradictory").sum()
    )
    if observed_forbidden or contradictory_rows:
        raise RuntimeError(
            "Post-review final384 violates fail-closed delivery: "
            f"forbidden_queues={observed_forbidden}, contradictory_rows={contradictory_rows}"
        )

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    review_disposition = pd.DataFrame()
    if args.selected384:
        source_disposition = Path(args.selected384).with_name(
            "agent_review_pool_post_review_disposition_v4_complete.csv"
        )
        if not source_disposition.exists():
            raise FileNotFoundError(source_disposition)
        review_disposition = pd.read_csv(source_disposition, low_memory=False).fillna("")
        if len(review_disposition) != int(formal_summary.get("agent_review_pool_rows", 512)):
            raise RuntimeError("Review-pool disposition row contract failed")
        if sha256(source_disposition) != str(
            selected_summary.get("output_sha256", {}).get("review_disposition", "")
        ):
            raise RuntimeError("Review-pool disposition SHA-256 differs from selection summary")
    final384_full_path = out_dir / "FINAL384_REVIEWED_FULL_V4.csv"
    final384_zh_path = out_dir / "FINAL384_REVIEWED_TEACHER_ZH_V4.csv"
    final1000_path_out = out_dir / "FINAL1000_RESERVE_FULL_V4.csv"
    review_disposition_path = out_dir / "REVIEW512_POST_REVIEW_DISPOSITION_FULL_V4.csv"
    final384.to_csv(final384_full_path, index=False)
    chinese = chinese_table(final384)
    chinese.to_csv(final384_zh_path, index=False)
    final1000.to_csv(final1000_path_out, index=False)
    if not review_disposition.empty:
        review_disposition.to_csv(review_disposition_path, index=False)
    known_controls = pd.read_csv(args.known_controls, low_memory=False).fillna("")
    mechanisms = pd.read_csv(args.mechanisms, low_memory=False).fillna("")
    drug_library = pd.read_csv(args.drug_library, low_memory=False).fillna("")
    assay_matrix = build_assay_matrix(final384, known_controls, mechanisms, drug_library)
    assay_matrix_path = out_dir / "FINAL384_ASSAY_MATRIX_V4.csv"
    assay_matrix.to_csv(assay_matrix_path, index=False)
    plate_map = build_nomination_plate_map(final384)
    plate_map_path = out_dir / "FINAL384_FOUR_BY_96_NOMINATION_MAP_V4.csv"
    plate_map.to_csv(plate_map_path, index=False)

    workbook_path = out_dir / "FDA_OLD_DRUG_NEW_TARGET_FINAL_PACKAGE_V4.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        chinese.to_excel(writer, index=False, sheet_name="384_中文审阅")
        final384.to_excel(writer, index=False, sheet_name="384_完整字段")
        final1000.to_excel(writer, index=False, sheet_name="1000_储备池")
        if not review_disposition.empty:
            review_disposition.to_excel(writer, index=False, sheet_name="512_审阅处置")
        final384[final384["review_queue"].eq("A_novel_priority")].to_excel(
            writer, index=False, sheet_name="A_新颖优先"
        )
        final384[final384["review_queue"].eq("P_validated_control_or_rediscovery")].to_excel(
            writer, index=False, sheet_name="P_对照再发现"
        )
        final384.groupby(["target_assay_family_v2", "primary_gene"], dropna=False).size().rename("pair_count").reset_index().to_excel(
            writer, index=False, sheet_name="实验分组"
        )
        assay_matrix.to_excel(writer, index=False, sheet_name="384_实验矩阵")
        plate_map.to_excel(writer, index=False, sheet_name="4x96_候选位图")
    style_excel(workbook_path)

    funnel = read_json(Path(args.funnel_summary))
    formal = formal_summary
    known = read_json(Path(args.known_summary))
    boltz_known = read_json(Path(args.boltz_known_summary)) if args.boltz_known_summary else {}
    rank_sensitivity = (
        read_json(Path(args.rank_sensitivity_summary)) if args.rank_sensitivity_summary else {}
    )
    pdf_path = out_dir / "FDA_OLD_DRUG_NEW_TARGET_FULL_REPORT_ZH_V4.pdf"
    build_pdf(
        pdf_path,
        final1000,
        final384,
        funnel,
        formal,
        known,
        boltz_known,
        rank_sensitivity,
    )
    cards_pdf_path = out_dir / "FDA_OLD_DRUG_NEW_TARGET_384_DETAILED_CARDS_ZH_V4.pdf"
    build_detailed_cards_pdf(cards_pdf_path, final384)

    files = [
        final384_full_path,
        final384_zh_path,
        final1000_path_out,
        assay_matrix_path,
        plate_map_path,
        workbook_path,
        pdf_path,
        cards_pdf_path,
    ]
    if not review_disposition.empty:
        files.append(review_disposition_path)
    manifest = {
        "created_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status": "formal_reviewed_delivery",
        "final1000_rows": len(final1000),
        "final384_rows": len(final384),
        "final384_unique_pairs": int(final384["pair_id"].nunique()),
        "review_pool_rows": int(len(review_disposition)) if not review_disposition.empty else 384,
        "review_queue_counts": final384["review_queue"].value_counts().to_dict(),
        "agent_grade_counts": final384["agent_feasibility_grade"].value_counts().to_dict(),
        "chembl_exact_binding_pchembl_ge_5_rows": int(
            final384.get("chembl_exact_activity_status", pd.Series("", index=final384.index))
            .astype(str)
            .eq("exact_binding_activity_pchembl_ge_5")
            .sum()
        ),
        "agent_exact_pair_validated_rows": int(
            final384["agent_literature_class"].astype(str).eq("exact_pair_validated").sum()
        ),
        "validated_control_or_rediscovery_union_rows": int(
            final384["review_queue"].eq("P_validated_control_or_rediscovery").sum()
        ),
        "database_query_incomplete_rows": int(
            final384["review_queue"].eq("Q_database_query_incomplete").sum()
        ),
        "hot_target_rows": int(bool_series(final384, "is_hot_target_2026").sum()),
        "hot_target_unique_genes": int(
            final384.loc[bool_series(final384, "is_hot_target_2026"), "primary_gene"].nunique()
        ),
        "files": {path.name: {"sha256": sha256(path), "bytes": path.stat().st_size} for path in files},
    }
    input_paths = {
        "formal_final1000": final1000_path,
        "formal_summary": formal_summary_path,
        "reviewed_or_selected384": Path(args.selected384 or args.reviewed384),
        "opentargets": Path(args.opentargets_final1000) if args.opentargets_final1000 else None,
        "funnel_summary": Path(args.funnel_summary),
        "known_summary": Path(args.known_summary),
        "boltz_known_summary": Path(args.boltz_known_summary) if args.boltz_known_summary else None,
        "known_controls": Path(args.known_controls),
        "mechanisms": Path(args.mechanisms),
        "drug_library": Path(args.drug_library),
        "hot_targets": Path(args.hot_targets),
        "config": Path(args.config),
        "chembl_cache": Path(args.chembl_cache) if args.chembl_cache else None,
        "pubmed_cache": Path(args.pubmed_cache) if args.pubmed_cache else None,
        "rank_sensitivity_summary": Path(args.rank_sensitivity_summary)
        if args.rank_sensitivity_summary
        else None,
        "sequence_audit_summary": Path(args.sequence_audit_summary)
        if args.sequence_audit_summary
        else None,
    }
    if args.selected384:
        input_paths["selected384_summary"] = Path(args.selected384).with_suffix(".summary.json")
        input_paths["review_pool_disposition_source"] = Path(args.selected384).with_name(
            "agent_review_pool_post_review_disposition_v4_complete.csv"
        )
    input_paths["boltz_run_plan"] = Path(
        "outputs/current_production_package_v2/full_untruncated_universe_v4/"
        "boltz_full_run_v4_seeded/run_plan.json"
    )
    input_paths["boltz_output_provenance"] = Path(
        "outputs/current_production_package_v2/full_untruncated_universe_v4/"
        "boltz_full_run_v4_seeded/result_provenance_with_output_hashes.csv"
    )
    input_paths["pyproject"] = Path("pyproject.toml")
    missing_inputs = [str(path) for path in input_paths.values() if path is not None and not path.exists()]
    if missing_inputs:
        raise FileNotFoundError(f"Delivery provenance inputs are missing: {missing_inputs}")
    config = yaml.safe_load(Path(args.config).read_text(encoding="utf-8"))
    for label, config_key, hash_key in [
        ("project_drug_manifest", "project_drug_manifest", "project_drug_manifest_sha256"),
        ("project_target_manifest", "project_target_manifest", "project_target_manifest_sha256"),
        (
            "project_target_integrity_manifest",
            "project_target_integrity_manifest",
            "project_target_integrity_manifest_sha256",
        ),
    ]:
        path = Path(config["inputs"][config_key])
        if sha256(path) != str(config["scope"][hash_key]):
            raise RuntimeError(f"Frozen manifest SHA-256 mismatch: {label}")
        input_paths[label] = path
    sequence_path = Path(config["inputs"]["protein_sequence_representatives"])
    if sha256(sequence_path) != str(config["scope"]["protein_sequence_representatives_sha256"]):
        raise RuntimeError("Frozen protein sequence table SHA-256 mismatch")
    input_paths["protein_sequence_representatives"] = sequence_path
    manifest["input_files"] = {
        label: {"path": str(path), "sha256": sha256(path), "bytes": path.stat().st_size}
        for label, path in input_paths.items()
        if path is not None
    }
    code_paths = [
        Path("biomaster/production.py"),
        Path("scripts/build_full_project_universe_v3.py"),
        Path("scripts/build_full_project_universe_v4.py"),
        Path("scripts/build_boltz2_complex_input_package.py"),
        Path("scripts/run_boltz2_batched_queue.py"),
        Path("scripts/recover_boltz2_missing_rows_v4.py"),
        Path("scripts/rebuild_boltz_output_provenance_v4.py"),
        Path("scripts/finalize_boltz_refined_3000_package.py"),
        Path("scripts/audit_boltz_pose_stability.py"),
        Path("scripts/finalize_106k_reselection_v2.py"),
        Path("scripts/finalize_full_universe_v4.py"),
        Path("scripts/finalize_known_control_boltz96_v4.py"),
        Path("scripts/audit_boltz_known_positive_calibration_v4.py"),
        Path("scripts/audit_final_candidates_chembl_activity.py"),
        Path("scripts/audit_final_candidates_pubmed.py"),
        Path("scripts/build_comprehensive_repurposing_literature_report.py"),
        Path("scripts/prepare_final384_agent_review_batches.py"),
        Path("scripts/merge_final384_agent_reviews.py"),
        Path("scripts/select_reviewed_final384_v4.py"),
        Path("scripts/build_final_v4_delivery_package.py"),
        Path("scripts/audit_final_delivery_v4.py"),
    ]
    tracked_code_paths = [str(path) for path in code_paths if path.exists()]
    dirty_formal = subprocess.run(
        ["git", "diff", "--quiet", "HEAD", "--", *tracked_code_paths],
        check=False,
    ).returncode
    if dirty_formal != 0:
        raise RuntimeError("Formal delivery source files differ from the current Git commit")
    git_commit = subprocess.check_output(["git", "rev-parse", "HEAD"], text=True).strip()
    git_tags = subprocess.check_output(
        ["git", "tag", "--points-at", "HEAD"], text=True
    ).splitlines()
    manifest["code_files"] = {
        str(path): sha256(path) for path in code_paths if path.exists()
    }
    manifest["source_control"] = {
        "git_commit": git_commit,
        "git_tags_at_commit": sorted(git_tags),
        "formal_source_files_clean_against_head": True,
    }
    (out_dir / "FINAL_DELIVERY_MANIFEST_V4.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
