#!/usr/bin/env python3
"""Fail-closed audit for the complete v4 CSV/XLSX/PDF delivery directory."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def truthy(series: pd.Series) -> pd.Series:
    return series.astype(str).str.lower().isin({"true", "1", "1.0", "yes"})


def counts(frame: pd.DataFrame, column: str, fallback: str = "") -> Counter:
    values = frame[column].astype(str) if column in frame else pd.Series(fallback, index=frame.index)
    return Counter(values)


def assert_caps(frame: pd.DataFrame, config: dict) -> dict[str, int]:
    cfg = config["selection"]["final384"]
    active = frame["active_moiety_smiles"].astype(str).where(
        frame["active_moiety_smiles"].astype(str).ne(""), frame["drug_chembl_id"].astype(str)
    )
    scaffold = frame["murcko_scaffold"].astype(str).where(
        frame["murcko_scaffold"].astype(str).ne(""), active
    )
    maxima = {
        "drug": int(active.value_counts().max()),
        "target": int(frame["primary_gene"].astype(str).value_counts().max()),
        "scaffold": int(scaffold.value_counts().max()),
    }
    limits = {
        "drug": int(cfg["drug_cap"]),
        "target": int(cfg["target_cap"]),
        "scaffold": int(cfg["scaffold_cap"]),
    }
    for key in maxima:
        if maxima[key] > limits[key]:
            raise RuntimeError(f"Final384 {key} cap failed: {maxima[key]} > {limits[key]}")
    families = frame["target_assay_family_v2"].value_counts().to_dict()
    for family, limit in cfg["family_caps"].items():
        if int(families.get(family, 0)) > int(limit):
            raise RuntimeError(
                f"Final384 family cap failed: {family}={families.get(family, 0)} > {limit}"
            )
    return maxima


def pdf_audit(path: Path, expected_pages: int | None = None) -> dict[str, int]:
    import fitz

    document = fitz.open(path)
    pages = len(document)
    if expected_pages is not None and pages != expected_pages:
        raise RuntimeError(f"PDF page count failed for {path.name}: {pages} != {expected_pages}")
    text_pages = sum(bool(page.get_text().strip()) for page in document)
    if text_pages != pages:
        raise RuntimeError(f"PDF contains blank/textless pages: {path.name} {text_pages}/{pages}")
    return {"pages": pages, "text_pages": text_pages}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default="configs/current_pipeline_v4.yaml")
    parser.add_argument(
        "--delivery-dir", default="outputs/current_production_package_v2/final_delivery_v4"
    )
    args = parser.parse_args()
    delivery = Path(args.delivery_dir)
    config = yaml.safe_load(Path(args.config).read_text(encoding="utf-8"))
    paths = {
        "final384": delivery / "FINAL384_REVIEWED_FULL_V4.csv",
        "final1000": delivery / "FINAL1000_RESERVE_FULL_V4.csv",
        "review512": delivery / "REVIEW512_POST_REVIEW_DISPOSITION_FULL_V4.csv",
        "assay": delivery / "FINAL384_ASSAY_MATRIX_V4.csv",
        "plate": delivery / "FINAL384_FOUR_BY_96_NOMINATION_MAP_V4.csv",
        "xlsx": delivery / "FDA_OLD_DRUG_NEW_TARGET_FINAL_PACKAGE_V4.xlsx",
        "main_pdf": delivery / "FDA_OLD_DRUG_NEW_TARGET_FULL_REPORT_ZH_V4.pdf",
        "cards_pdf": delivery / "FDA_OLD_DRUG_NEW_TARGET_384_DETAILED_CARDS_ZH_V4.pdf",
        "manifest": delivery / "FINAL_DELIVERY_MANIFEST_V4.json",
    }
    missing = [str(path) for path in paths.values() if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Final delivery files are missing: {missing}")
    final384 = pd.read_csv(paths["final384"], low_memory=False).fillna("")
    final1000 = pd.read_csv(paths["final1000"], low_memory=False).fillna("")
    review512 = pd.read_csv(paths["review512"], low_memory=False).fillna("")
    assay = pd.read_csv(paths["assay"], low_memory=False).fillna("")
    plate = pd.read_csv(paths["plate"], low_memory=False).fillna("")
    for label, frame, expected in [
        ("final384", final384, 384),
        ("final1000", final1000, 1000),
        ("review512", review512, 512),
        ("assay", assay, 384),
        ("plate", plate, 384),
    ]:
        if len(frame) != expected:
            raise RuntimeError(f"{label} row contract failed: {len(frame)} != {expected}")
    if final384["pair_id"].duplicated().any() or final1000["pair_id"].duplicated().any():
        raise RuntimeError("Final package contains duplicate pair_id values")
    if not set(final384["pair_id"]).issubset(set(final1000["pair_id"])):
        raise RuntimeError("Final384 is not nested in final1000")
    forbidden_boolean = [
        "exact_known_target_v2",
        "family_or_rediscovery_risk_v2",
        "severe_compound_liability",
        "ion_channel_feasibility_flag",
        "structure_sequence_mismatch_v4",
    ]
    for column in forbidden_boolean:
        if truthy(final384[column]).any():
            raise RuntimeError(f"Final384 contains forbidden rows: {column}")
    if final384["agent_feasibility_grade"].eq("D").any():
        raise RuntimeError("Final384 contains grade-D rows")
    if final384["agent_literature_class"].eq("contradictory").any():
        raise RuntimeError("Final384 contains contradictory rows")
    if final384["review_queue"].isin(
        ["D_deprioritize", "Q_database_query_incomplete"]
    ).any():
        raise RuntimeError("Final384 contains fail-closed review queues")
    if not final384["pose_stability_tier"].isin(
        ["A_stable_conditional_pose", "B_moderate_conditional_pose"]
    ).all():
        raise RuntimeError("Final384 contains unsupported conditional poses")
    control_rows = int(
        final384["review_candidate_class_v4"].eq("validated_control_or_rediscovery").sum()
    )
    control_cap = int(config["selection"]["post_review"]["validated_control_cap"])
    if control_rows > control_cap:
        raise RuntimeError(f"Validated-control cap failed: {control_rows} > {control_cap}")
    required_agent = [
        "agent_verdict",
        "agent_literature_class",
        "agent_primary_disease",
        "agent_repurposing_status",
        "agent_disease_evidence",
        "agent_mechanism_rationale",
        "agent_exposure_feasibility",
        "agent_active_species_status",
        "agent_assay_plan",
        "agent_key_risks",
        "agent_database_query_resolution",
        "agent_sources",
    ]
    missing_agent = {
        column: int(final384[column].astype(str).str.strip().eq("").sum())
        for column in required_agent
        if final384[column].astype(str).str.strip().eq("").any()
    }
    if missing_agent:
        raise RuntimeError(f"Final384 agent fields are incomplete: {missing_agent}")
    if final384["agent_active_species_status"].eq(
        "prodrug_active_metabolite_requires_rerun"
    ).any():
        raise RuntimeError("Final384 contains a prodrug requiring active-metabolite rerun")
    maxima = assert_caps(final384, config)
    block_sizes = plate.groupby("nomination_block").size().to_dict()
    if block_sizes != {f"Block_{index}": 96 for index in range(1, 5)}:
        raise RuntimeError(f"4x96 nomination map failed: {block_sizes}")
    if assay["pair_id"].nunique() != 384 or plate["pair_id"].nunique() != 384:
        raise RuntimeError("Assay or nomination map does not cover 384 unique pairs")

    workbook = load_workbook(paths["xlsx"], read_only=True)
    expected_sheets = {
        "384_中文审阅",
        "384_完整字段",
        "1000_储备池",
        "512_审阅处置",
        "384_实验矩阵",
        "4x96_候选位图",
    }
    if not expected_sheets.issubset(set(workbook.sheetnames)):
        raise RuntimeError(f"Workbook sheets are missing: {sorted(expected_sheets - set(workbook.sheetnames))}")
    main_pdf = pdf_audit(paths["main_pdf"])
    cards_pdf = pdf_audit(paths["cards_pdf"], expected_pages=96)
    manifest = json.loads(paths["manifest"].read_text(encoding="utf-8"))
    for name, metadata in manifest.get("files", {}).items():
        path = delivery / name
        if not path.exists() or sha(path) != metadata.get("sha256"):
            raise RuntimeError(f"Delivery manifest hash failed: {name}")
    result = {
        "status": "formal_delivery_audit_passed",
        "rows": {key: len(frame) for key, frame in {
            "final384": final384,
            "final1000": final1000,
            "review512": review512,
            "assay": assay,
            "plate": plate,
        }.items()},
        "final384_maxima": maxima,
        "final384_family_counts": dict(counts(final384, "target_assay_family_v2")),
        "validated_control_rows": control_rows,
        "plate_block_sizes": block_sizes,
        "main_pdf": main_pdf,
        "cards_pdf": cards_pdf,
        "manifest_sha256": sha(paths["manifest"]),
    }
    output = delivery / "FINAL_DELIVERY_AUDIT_V4.json"
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
