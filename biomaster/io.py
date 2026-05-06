from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable


def ensure_parent(path: str | Path) -> Path:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def read_csv_rows(path: str | Path) -> list[dict[str, str]]:
    with Path(path).open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def read_table_rows(path: str | Path, sheet: str | int | None = None) -> list[dict[str, str]]:
    target = Path(path)
    suffix = target.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(target)
    if suffix == ".tsv":
        with target.open(newline="", encoding="utf-8") as handle:
            return list(csv.DictReader(handle, delimiter="\t"))
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(target, sheet=sheet)
    raise ValueError(f"Unsupported table format for {target}; expected .csv, .tsv, .xlsx, or .xlsm")


def read_xlsx_rows(path: Path, sheet: str | int | None = None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("Reading .xlsx files requires openpyxl. Install it with: pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if isinstance(sheet, str) else workbook.worksheets[sheet or 0]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(rows)
    except StopIteration:
        return []

    headers = ["" if value is None else str(value).strip() for value in header_values]
    output: list[dict[str, str]] = []
    for values in rows:
        row: dict[str, str] = {}
        has_value = False
        for header, value in zip(headers, values):
            if not header:
                continue
            text = "" if value is None else str(value).strip()
            if text:
                has_value = True
            row[header] = text
        if has_value:
            output.append(row)
    return output


def write_csv_rows(path: str | Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    target = ensure_parent(path)
    with target.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: "" if row.get(key) is None else row.get(key) for key in fieldnames})


def as_float(value: object, default: float | None = None) -> float | None:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def as_int(value: object, default: int | None = None) -> int | None:
    number = as_float(value, None)
    if number is None:
        return default
    return int(number)


def first_present(row: dict[str, object], keys: Iterable[str], default: str = "") -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return default
